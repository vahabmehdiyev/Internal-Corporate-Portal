import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from api.graphs_reports.graphs_reports_dao import get_psi_data_by_line_and_date, get_asse_data_by_line_and_date, get_hours_data_by_line_and_date
from openpyxl.chart import BarChart, LineChart, Reference


def get_unique_filename(folder: str, base_name: str, extension: str = ".xlsx") -> tuple[str, str]:
    count = 0
    filename = f"{base_name}{extension}"
    file_path = os.path.join(folder, filename)

    while os.path.exists(file_path):
        count += 1
        filename = f"{base_name} ({count}){extension}"
        file_path = os.path.join(folder, filename)

    return file_path, filename

def build_table_from_rows(rows, product_key="name"):
    # Step 1: Saatları və məhsul adlarını topla
    hours = sorted(set(row["hour"] for row in rows))
    product_names = sorted(set(row[product_key] for row in rows))

    # Step 2: Data strukturunu qur
    table_data = {name: {hour: 0 for hour in hours} for name in product_names}

    for row in rows:
        name = row[product_key]
        hour = row["hour"]
        total = row["total"]
        table_data[name][hour] += total

    # Step 3: Header
    headers = ["Наименование изделия"] + [f"{h} час" for h in hours] + ["Итого"]

    # Step 4: Row-ları yığ
    table_rows = []
    for name, hour_data in table_data.items():
        row = [name]
        row_total = 0
        for h in hours:
            val = hour_data[h]
            row.append(val)
            row_total += val
        row.append(row_total)
        table_rows.append(row)

    # Step 5: Footer (sütunların cəmi)
    footer = ["Итого по часам"]
    for h in hours:
        col_sum = sum(table_data[name][h] for name in product_names)
        footer.append(col_sum)
    footer.append(sum(footer[1:]))  # Total sum
    table_rows.append(footer)

    return {
        "headers": headers,
        "rows": table_rows
    }

def get_table_for_excel_export(data):
    source = data.get("source")
    date = data.get("date")
    line = data.get("line")

    #Xam datanı çək və konfiqurasiya et
    if source == "psi":
        rows = get_psi_data_by_line_and_date(data)
        product_key = "comments"
        title_prefix = "Таблица проверенных изделий"
    elif source == "asse":
        rows = get_asse_data_by_line_and_date(data)
        product_key = "name"
        title_prefix = "Таблица собранных изделий"
    elif source == "hours":
        rows = get_hours_data_by_line_and_date(data)
        product_key = "comments"
        title_prefix = "Таблица упакованных изделий"
    else:
        raise ValueError(f"Naməlum mənbə: {source}")

    table = build_table_from_rows(rows, product_key=product_key)

    # Fayl adı və yolu
    base_name = f"{title_prefix} за {date} {line} конвейера"
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    file_path, _ = get_unique_filename(desktop_path, base_name)

    # Excel workbook və sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Таблица"

    # Format stilləri
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    #  Başlıq: Конвейер X
    total_columns = len(table["headers"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    cell = ws.cell(row=1, column=1, value=f"Конвейер {line}")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(size=12, bold=True)

    #  Header sətiri
    for col_idx, val in enumerate(table["headers"], start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.font = Font(bold=True)
        cell.border = thin_border

    #  Cədvəl satırları
    for row_idx, row in enumerate(table["rows"], start=3):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True if col_idx == 1 else False  #  Məhsul adı (1-ci sütun) üçün wrap text aktiv
            )
            cell.border = thin_border

    #  Gridlines gizlət (çərçivələr var artıq)
    ws.sheet_view.showGridLines = False

    #  Sütun genişliyi: 1-ci sütun geniş, digərləri dar
    for col_idx in range(1, total_columns + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx == 1:
            ws.column_dimensions[col_letter].width = 25  # Məhsul adı sütunu
        else:
            ws.column_dimensions[col_letter].width = 10  # Saat sütunları

    #  Faylı yadda saxla
    wb.save(file_path)
    print(f" Formatlı Excel faylı yaradıldı: {file_path}")
    return file_path


def get_graph_for_excel_export(data):
    source = data.get("source")
    date = data.get("date")
    line = data.get("line")

    #  Raw data al
    if source == "psi":
        rows = get_psi_data_by_line_and_date(data)
        title_prefix = "График проверенных изделий"
        product_key = "comments"
    elif source == "asse":
        rows = get_asse_data_by_line_and_date(data)
        title_prefix = "График собранных изделий"
        product_key = "name"
    elif source == "hours":
        rows = get_hours_data_by_line_and_date(data)
        title_prefix = "График упакованных изделий"
        product_key = "comments"
    else:
        raise ValueError(f"Naməlum mənbə: {source}")

    #  Məhsul adları və saatlar
    product_names = sorted(set(row[product_key] for row in rows))
    all_hours = sorted(set(row["hour"] for row in rows))

    # Saatlara görə məlumat xəritəsi
    hour_map = {hour: {name: 0 for name in product_names} for hour in all_hours}
    for row in rows:
        name = row[product_key]
        hour = row["hour"]
        total = row["total"]
        hour_map[hour][name] += total

    #  DataFrame hazırla
    graph_data = []
    for hour in all_hours:
        item = {name: hour_map[hour][name] for name in product_names}
        item["count"] = sum(item.values())
        item["label"] = f"{hour:02}:00"
        graph_data.append(item)

    df = pd.DataFrame(graph_data)

    #  Fayl yolu
    base_name = f"{title_prefix} за {date} {line} конвейера"
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    file_path = os.path.join(desktop_path, base_name + ".xlsx")
    if os.path.exists(file_path):
        i = 1
        while os.path.exists(f"{file_path[:-5]} ({i}).xlsx"):
            i += 1
        file_path = f"{file_path[:-5]} ({i}).xlsx"

    #  Excel faylı və sheet-lər
    wb = Workbook()
    data_ws = wb.active
    data_ws.title = "Data"
    chart_ws = wb.create_sheet(title="Graph")

    for col_idx, col in enumerate(df.columns, start=1):
        data_ws.cell(row=1, column=col_idx, value=col)
    for row_idx, row in enumerate(df.values, start=2):
        for col_idx, val in enumerate(row, start=1):
            data_ws.cell(row=row_idx, column=col_idx, value=val)

    data_ws.sheet_state = "hidden"

    #  Qrafik
    bar_chart = BarChart()
    bar_chart.title = f"{title_prefix} — Конвейер {line}"
    bar_chart.y_axis.title = ""
    bar_chart.x_axis.title = ""
    bar_chart.width = 25
    bar_chart.height = 15
    bar_chart.legend.position = 'b'
    bar_chart.title.font = Font(size=16, bold=True)
    bar_chart.x_axis.title.font = Font(size=12)
    bar_chart.y_axis.title.font = Font(size=12)
    bar_chart.x_axis.title.overlay = False  # "Часы" ilə üst-üstə düşməsin

    count_col_idx = df.columns.get_loc("count") + 1
    label_col_idx = df.columns.get_loc("label") + 1
    num_rows = len(df.index) + 1

    labels = Reference(data_ws, min_col=label_col_idx, min_row=2, max_row=num_rows)
    bar_data = Reference(data_ws, min_col=count_col_idx, min_row=1, max_row=num_rows)

    bar_chart.add_data(bar_data, titles_from_data=True)
    bar_chart.set_categories(labels)

    #  Bar rəngi dəyiş: count üçün
    bar_chart.series[0].graphicalProperties.solidFill = "93c5fd"  # Açıq mavi

    #  Məhsullar üçün LineChart
    line_chart = LineChart()
    for col in df.columns:
        if col not in ("label", "count"):
            idx = df.columns.get_loc(col) + 1
            line_data = Reference(data_ws, min_col=idx, min_row=1, max_row=num_rows)
            line_chart.add_data(line_data, titles_from_data=True)
    line_chart.set_categories(labels)

    bar_chart += line_chart
    chart_ws.add_chart(bar_chart, "B2")

    # 8️ Saxla və qaytar
    wb.save(file_path)
    print(f"✅ Qrafikli Excel faylı yaradıldı: {file_path}")
    return file_path




def build_combined_excel_file_all_lines(source, date):
    #  Mənbəyə görə konfiqurasiya
    if source == "psi":
        title_prefix = "Отчёт проверенных изделий"
        get_rows_fn = get_psi_data_by_line_and_date
        product_key = "comments"
    elif source == "asse":
        title_prefix = "Отчёт собранных изделий"
        get_rows_fn = get_asse_data_by_line_and_date
        product_key = "name"
    elif source == "hours":
        title_prefix = "Отчёт упакованных изделий"
        get_rows_fn = get_hours_data_by_line_and_date
        product_key = "comments"
    else:
        raise ValueError(f"Naməlum mənbə: {source}")

    wb = Workbook()
    wb.remove(wb.active)  # Default sheet sil

    for line in range(1, 8):
        data = {"date": date, "line": line}
        rows = get_rows_fn(data)

        if not rows:
            continue  # Bu line üçün data yoxdursa, keç

        table = build_table_from_rows(rows, product_key=product_key)
        ws = wb.create_sheet(title=f"Конвейер {line}")

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        total_columns = len(table["headers"])

        #  Başlıq
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
        cell = ws.cell(row=1, column=1, value=f"Конвейер {line}")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(size=12, bold=True)

        #  Header
        for col_idx, val in enumerate(table["headers"], start=1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True)
            cell.border = thin_border

        #  Data
        for row_idx, row in enumerate(table["rows"], start=3):
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True if col_idx == 1 else False
                )
                cell.border = thin_border

        ws.sheet_view.showGridLines = False

        for col_idx in range(1, total_columns + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 25 if col_idx == 1 else 10

        #  Qrafik üçün data hazırlanması
        hours = sorted(set(row["hour"] for row in rows))
        product_names = sorted(set(row[product_key] for row in rows))
        hour_map = {hour: {name: 0 for name in product_names} for hour in hours}

        for row in rows:
            name = row[product_key]
            hour = row["hour"]
            total = row["total"]
            hour_map[hour][name] += total

        graph_data = []
        for hour in hours:
            item = {name: hour_map[hour][name] for name in product_names}
            item["count"] = sum(item.values())
            item["label"] = f"{hour:02}:00"
            graph_data.append(item)

        if not graph_data:
            continue  # qrafik üçün data yoxdursa keç

        df = pd.DataFrame(graph_data)

        # Gizli Data sheet
        data_ws = wb.create_sheet(title=f"Data_{line}")
        for col_idx, col in enumerate(df.columns, start=1):
            data_ws.cell(row=1, column=col_idx, value=col)
        for row_idx, row in enumerate(df.values, start=2):
            for col_idx, val in enumerate(row, start=1):
                data_ws.cell(row=row_idx, column=col_idx, value=val)
        data_ws.sheet_state = "hidden"

        #  Qrafik
        bar_chart = BarChart()
        bar_chart.title = f"График за {date} — Конвейер {line}"
        bar_chart.width = 25
        bar_chart.height = 15
        bar_chart.legend.position = 'b'
        bar_chart.x_axis.title = ""
        bar_chart.y_axis.title = ""

        try:
            count_col_idx = df.columns.get_loc("count") + 1
            label_col_idx = df.columns.get_loc("label") + 1
        except KeyError as e:
            print(f"Qrafik üçün sütun tapılmadı: {e}")
            continue

        num_rows = len(df.index) + 1
        labels = Reference(data_ws, min_col=label_col_idx, min_row=2, max_row=num_rows)
        bar_data = Reference(data_ws, min_col=count_col_idx, min_row=1, max_row=num_rows)

        bar_chart.add_data(bar_data, titles_from_data=True)
        bar_chart.set_categories(labels)
        bar_chart.series[0].graphicalProperties.solidFill = "93c5fd"

        # Line chart (məhsullar üzrə)
        line_chart = LineChart()
        for col in df.columns:
            if col not in ("label", "count"):
                idx = df.columns.get_loc(col) + 1
                line_data = Reference(data_ws, min_col=idx, min_row=1, max_row=num_rows)
                line_chart.add_data(line_data, titles_from_data=True)
        line_chart.set_categories(labels)
        bar_chart += line_chart

        #Qrafiki yerləşdir
        graph_anchor_row = len(table["rows"]) + 5
        graph_anchor_cell = f"A{graph_anchor_row}"
        ws.add_chart(bar_chart, graph_anchor_cell)

    # Faylı saxla
    file_path = os.path.join(
        os.path.expanduser("~"),
        "Desktop",
        f"{title_prefix} за {date}.xlsx"
    )
    if os.path.exists(file_path):
        i = 1
        while os.path.exists(f"{file_path[:-5]} ({i}).xlsx"):
            i += 1
        file_path = f"{file_path[:-5]} ({i}).xlsx"

    wb.save(file_path)
    print(f"✅ Ümumi Excel faylı yaradıldı: {file_path}")
    return file_path